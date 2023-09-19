import pandas as pd
import pygame
import sys
from pygame.locals import *
from moviepy.editor import *
import openpyxl

# creating pandas dataframe from dictionary of data
df_personal = pd.DataFrame({'Employee Name': ['Guy'],
                        'Employee number': ['Ai001'],
                        'Title': ["Chief architect"],
                        'DOJ': ["1-Apr-2007"]})
# Exporting dataframe to Excel file
pygame.init()
clip = VideoFileClip("D:\\SHASHANK PATIL 12 A\\COMP PROJECT\\TAX-program__school-main\\vid.mp4").resize(0.65)

clip.preview(fps=20)
def inp():
    bg_img = pygame.image.load("D:\\SHASHANK PATIL 12 A\\COMP PROJECT\\TAX-program__school-main\\wallpaper.jpg")
    clock = pygame.time.Clock()
    screen = pygame.display.set_mode([1270,720])
    base_font = pygame.font.Font(None, 38)
    user_text = 'Enter the ctc:'
    user_text1 = 'Provide 80C exemption'
    user_text2 = 'Provide 80d exemption'
    user_text3 = 'provide rent'
    submit_text="SUBMIT"
    submit_rect = pygame.Rect(900, 600, 140, 38)
    input_rect = pygame.Rect(200, 200, 140, 38)
    input_rect1 = pygame.Rect(500, 200, 140, 38)
    input_rect2 = pygame.Rect(500, 325, 140, 38)
    input_rect3 = pygame.Rect(500, 450, 140, 38)
    color_active = (50,50,50)
    color_passive = (13,137,189)
    color = color_passive
    color1=color_passive
    active = False
    active1 = False
    active2 = False
    active3 = False
    #screen.fill((255, 255, 255))
    T=True
    while T:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()

            if event.type == pygame.MOUSEBUTTONDOWN:
                if input_rect.collidepoint(event.pos):
                    active = True
                    user_text = ""
                else:
                    active = False
                if input_rect1.collidepoint(event.pos):
                    active1 = True
                    user_text1 = ""
                else:
                    active1 = False
                if input_rect2.collidepoint(event.pos):
                    active2 = True
                    user_text2 = ""
                else:
                    active2 = False
                if input_rect3.collidepoint(event.pos):
                    active3 = True
                    user_text3 = ""
                else:
                    active3 = False
                if submit_rect.collidepoint(event.pos):
                    print("Be=reaking")
                    T=False
                    return(user_text,user_text1,user_text2,user_text3)
                    break

            if event.type == pygame.KEYDOWN:

                            # Check for backspace
                if event.key == pygame.K_RETURN and active == True:
                    print(user_text)
                    active == False
                if event.key == pygame.K_BACKSPACE and active == True:

                    user_text = user_text[:-1]

                            # formation
                if event.key != pygame.K_RETURN and active == True and event.key != pygame.K_BACKSPACE:
                    user_text += event.unicode


                if event.key == pygame.K_RETURN and active1 == True:
                    print(user_text1)
                if event.key == pygame.K_BACKSPACE and active1 == True:

                    user_text1 = user_text1[:-1]

                            # formation
                if event.key != pygame.K_RETURN and active1 == True and event.key != pygame.K_BACKSPACE:
                    user_text1 += event.unicode
                #######################################
                if event.key == pygame.K_RETURN and active2 == True:
                    print(user_text2)
                if event.key == pygame.K_BACKSPACE and active2 == True:

                    user_text2 = user_text2[:-1]

                            # formation
                if event.key != pygame.K_RETURN and active2 == True and event.key != pygame.K_BACKSPACE:
                    user_text2 += event.unicode
                ######################################
                if event.key == pygame.K_RETURN and active3 == True:
                    print(user_text3)
                if event.key == pygame.K_BACKSPACE and active3 == True:

                    user_text3 = user_text3[:-1]

                            # formation
                if event.key != pygame.K_RETURN and active3 == True and event.key != pygame.K_BACKSPACE:
                    user_text3 += event.unicode
        #screen.fill((255, 255, 255))

        if active:
            color = color_active
        else:
            color = color_passive
        if active1:
            color1 = color_active
        else:
            color1 = color_passive
        if active2:
            color2 = color_active
        else:
            color2 = color_passive
        if active3:
            color3 = color_active
        else:
            color3 = color_passive
        pygame.draw.rect(screen, color_passive, submit_rect)
        pygame.draw.rect(screen, color, input_rect)
        pygame.draw.rect(screen, color1, input_rect1)
        pygame.draw.rect(screen, color2, input_rect2)
        pygame.draw.rect(screen, color3, input_rect3)

        submit_surface = base_font.render(submit_text, True, (255, 255, 255))
        text_surface = base_font.render(user_text, True, (255, 255, 255))
        text_surface1 = base_font.render(user_text1, True, (255, 255, 255))
        text_surface2 = base_font.render(user_text2, True, (255, 255, 255))
        text_surface3 = base_font.render(user_text3, True, (255, 255, 255))

        screen.blit(submit_surface, (submit_rect.x+5, submit_rect.y+5))
        screen.blit(text_surface, (input_rect.x+5, input_rect.y+5))
        screen.blit(text_surface1, (input_rect1.x+5, input_rect1.y+5))
        screen.blit(text_surface2, (input_rect2.x + 5, input_rect2.y + 5))
        screen.blit(text_surface3, (input_rect3.x + 5, input_rect3.y + 5))

        input_rect.w = max(140, text_surface.get_width()+10)
        input_rect1.w = max(140, text_surface1.get_width()+10)
        input_rect2.w = max(140, text_surface2.get_width() + 10)
        input_rect3.w = max(140, text_surface3.get_width() + 10)

        pygame.display.flip()
        clock.tick(60)
        screen.blit(bg_img, (0, 0))
j=inp()
ctc=int(j[0])
C_80Exemption=int(j[1])
D_80Exemption=int(j[2])
rent=int(j[3])
print(df_personal)
#ctc=int(input("Enter ctc"))
basic = (50/100)*ctc
hra=(40/100)*basic
pf=(12/100)*basic
spal=ctc-(basic+hra+pf)
#income tax calc
#C_80Exemption=int(input("Enter your 80 C exemption:"))
if C_80Exemption >150000:
    print("Invalid")
#D_80Exemption=int(input("Enter your 80 D exemption:"))
if D_80Exemption > 25000:
    print("invalid")
#rent=int(input("Enter rent:"))
#
Anual_basic=basic#date
Anual_hra=hra
Anual_sp=spal
total_=Anual_sp+Anual_hra+Anual_basic
cal=rent-((10/100)*basic)
hra_detuction=min([hra,cal])
sec80=C_80Exemption+D_80Exemption
total_detuction=hra_detuction+sec80
#print("Total:",total_)
#print(total_detuction)
def old(gross,less_dens,s1,s2,s3,su):
    less_dens=-less_dens
    Net_inc=gross+less_dens
    print(Net_inc)
    if Net_inc < 250000:
        tax_rate=0
    if Net_inc > 1000000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 20 / 100
        s2 = tax * (1000000 - 500000)
        tax = 30 / 100
        s3 = tax * (Net_inc - 1000000)
        su = s1 + s2 + s3
    if Net_inc > 500000 and Net_inc < 1000000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 20 / 100
        s2 = tax * (Net_inc - 500000)
        su = s1 + s2
    if Net_inc > 250000 and Net_inc < 500000:
        tax = 5 / 100
        su = tax * (Net_inc - 250000)
    Tax_cess=su
    cess=(4/100)*Tax_cess
    IT=Tax_cess+cess
    recoverO=IT/12
    print(recoverO)#date
    return recoverO
def new(gross,s1,s2,s3,s4,s5,s6,su):
    Net_inc = gross
    if Net_inc > 1500000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 10 / 100
        s2 = tax * (250000)
        tax = 15 / 100
        s3 = tax * (250000)
        tax = 20 / 100
        s4 = tax * (250000)
        tax = 25 / 100
        s5 = tax * (250000)
        tax = 30 / 100
        s6 = tax * (Net_inc - 1500000)
        # print(s1,s2,s3,s4,s5,s6,sep="\n")
        su = s1 + s2 + s3 + s4 + s5 + s6
    if Net_inc > 500000 and Net_inc < 750000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 10 / 100
        s2 = tax * (Net_inc - 500000)
        su = s1 + s2
    if Net_inc > 750000 and Net_inc < 1000000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 10 / 100
        s2 = tax * (250000)
        tax = 15 / 100
        s3 = tax * (Net_inc - 750000)
        su = s1 + s2 + s3
    if Net_inc > 1000000 and Net_inc < 1250000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 10 / 100
        s2 = tax * (250000)
        tax = 15 / 100
        s3 = tax * (250000)
        tax = 20 / 100
        s4 = tax * (Net_inc - 1000000)
        su = s1 + s2 + s3 + s4
    if Net_inc > 1250000 and Net_inc < 1500000:
        tax = 5 / 100
        s1 = tax * (500000 - 250000)
        tax = 10 / 100
        s2 = tax * (250000)
        tax = 15 / 100
        s3 = tax * (250000)
        tax = 20 / 100
        s4 = tax * (250000)
        tax = 25 / 100
        s5 = tax * (Net_inc - 1250000)
        su = s1 + s2 + s3 + s4 + s5
    if Net_inc > 250000 and Net_inc < 500000:
        tax = 5 / 100
        su = tax * (Net_inc- 250000)
    Tax_cess = su
    cess = (4 / 100) * Tax_cess
    IT = Tax_cess + cess
    recoverN = IT / 12
    print(recoverN)# date
    return recoverN
o=old(total_,total_detuction,0,0,0,0)
n=new(total_,0,0,0,0,0,0,0)
if o >n:
    IT=n
else:
    IT=o
te=(hra/12+basic/12+spal/12)
print(hra,basic,spal,te)
print(pf,IT)
td=pf/12+IT
print("net:", te-td)
#exel conversion
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
wb = Workbook()
ws = wb.active
ws. merge_cells('A1:D1')
cell = ws. cell(row=1, column=1)
cell. alignment = Alignment(horizontal='center', vertical='center')
ws['A1']="ABC Company Private Limited"
ws. merge_cells('F3:J3')
cell = ws. cell(row=3, column=6)
cell. alignment = Alignment(horizontal='center', vertical='center')
ws['F1']="Payroll of month"
for r in dataframe_to_rows(df_personal, index=False, header=True):
    ws.append(r)
ws['F4'] ="Earnings"
ws['G4'] = "Amount"
ws['I4'] ="Deduction"
ws['J4'] ="Amount"
ws['I5'] ="PF"
ws['J5'] =pf/12
ws['I6'] ="IT"
ws['J6'] =IT/12
ws['I8'] ="Total Deductions"
ws['J8'] =td
ws['F9'] ="NET PAY"
ws['G9']=te-td
ws['G8'] =te
ws['F8'] ="Total Earnings"
ws['G7'] =spal/12
ws['F7'] ="Special Allow"
ws['G6'] =hra/12
ws['F6'] ="HRA"
ws['G5'] = basic/12
ws['F5'] ="Basic"
wb.save("Tax.xlsx")
