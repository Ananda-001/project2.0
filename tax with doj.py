import pandas as pd
import pygame
import sys
from pygame.locals import *
from moviepy.editor import *
import openpyxl

df_personal = pd.DataFrame({'Employee Name': ['Guy'],
                            'Employee number': ['Ai001'],
                            'Title': ["Chief architect"],
                            'DOJ': ["1-Apr-2007"]})


def tax_calc():  # put in the values of j in here
    j = (40000, 8000, 2000, 10000)
    ctc = int(j[0])
    C_80Exemption = int(j[1])
    D_80Exemption = int(j[2])
    rent = int(j[3])
    # ctc=int(input("Enter the ctc"))
    # C_80Exemption=int(input("Enter the 80 C"))
    # D_80Exemption=int(input("ENter the 80 D"))
    # rent=int(input("Enter the rent:"))
    basic = (50 / 100) * ctc
    hra = (40 / 100) * basic
    pf = (12 / 100) * basic
    spal = ctc - (basic + hra + pf)
    # income tax calc
    if C_80Exemption > 150000:
        print("Invalid")
    if D_80Exemption > 25000:
        print("invalid")
    Anual_basic = basic  # date
    Anual_hra = hra
    Anual_sp = spal
    total_ = Anual_sp + Anual_hra + Anual_basic
    cal = rent - ((10 / 100) * basic)
    hra_detuction = min([hra, cal])
    sec80 = C_80Exemption + D_80Exemption
    total_detuction = hra_detuction + sec80

    def old(gross, less_dens, s1, s2, s3, su):
        less_dens = -less_dens
        Net_inc = gross + less_dens
        print(Net_inc)
        if Net_inc < 250000:
            tax_rate = 0
        if Net_inc > 1000000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 20 / 100
            s2 = tax * (1000000 - 500000)
            tax = 30 / 100
            s3 = tax * (Net_inc - 1000000)
            su = s1 + s2 + s3
        if Net_inc > 500000 and Net_inc < 1000000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 20 / 100
            s2 = tax * (Net_inc - 500000)
            su = s1 + s2
        if Net_inc > 250000 and Net_inc < 500000:
            tax = 5 / 100
            su = tax * (Net_inc - 250000)
        Tax_cess = su
        cess = (4 / 100) * Tax_cess
        IT = Tax_cess + cess
        recoverO = IT / 12
        print(recoverO)  # date
        return recoverO

    def new(gross, s1, s2, s3, s4, s5, s6, su):
        Net_inc = gross
        if Net_inc > 1500000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 10 / 100
            s2 = tax * (250000)
            tax = 15 / 100
            s3 = tax * (250000)
            tax = 20 / 100
            s4 = tax * (250000)
            tax = 25 / 100
            s5 = tax * (250000)
            tax = 30 / 100
            s6 = tax * (Net_inc - 1500000)
            su = s1 + s2 + s3 + s4 + s5 + s6
        if Net_inc > 500000 and Net_inc < 750000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 10 / 100
            s2 = tax * (Net_inc - 500000)
            su = s1 + s2
        if Net_inc > 750000 and Net_inc < 1000000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 10 / 100
            s2 = tax * (250000)
            tax = 15 / 100
            s3 = tax * (Net_inc - 750000)
            su = s1 + s2 + s3
        if Net_inc > 1000000 and Net_inc < 1250000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 10 / 100
            s2 = tax * (250000)
            tax = 15 / 100
            s3 = tax * (250000)
            tax = 20 / 100
            s4 = tax * (Net_inc - 1000000)
            su = s1 + s2 + s3 + s4
        if Net_inc > 1250000 and Net_inc < 1500000:
            tax = 5 / 100
            s1 = tax * (500000 - 250000)
            tax = 10 / 100
            s2 = tax * (250000)
            tax = 15 / 100
            s3 = tax * (250000)
            tax = 20 / 100
            s4 = tax * (250000)
            tax = 25 / 100
            s5 = tax * (Net_inc - 1250000)
            su = s1 + s2 + s3 + s4 + s5
        if Net_inc > 250000 and Net_inc < 500000:
            tax = 5 / 100
            su = tax * (Net_inc - 250000)
        Tax_cess = su
        cess = (4 / 100) * Tax_cess
        IT = Tax_cess + cess
        recoverN = IT / 12
        print(recoverN)  # date
        return recoverN

    o = old(total_, total_detuction, 0, 0, 0, 0)
    n = new(total_, 0, 0, 0, 0, 0, 0, 0)
    if o > n:
        used = "new"
        IT = n
    else:
        IT = o
        used = "old"
    te = (hra / 12 + basic / 12 + spal / 12)
    print(hra, basic, spal, te)
    print(pf, IT)
    td = pf / 12 + IT
    print("net:", te - td)
    print("used:", used)
    # exel conversion
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'] = "ABC Company Private Limited"
    ws.merge_cells('F3:J3')
    cell = ws.cell(row=3, column=6)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws['F1'] = "Payroll of month"
    for r in dataframe_to_rows(df_personal, index=False, header=True):
        ws.append(r)
    ws['F4'] = "Earnings"
    ws['G4'] = "Amount"
    ws['I4'] = "Deduction"
    ws['J4'] = "Amount"
    ws['I5'] = "PF"
    ws['J5'] = pf / 12
    ws['I6'] = "IT"
    ws['J6'] = IT / 12
    ws['I8'] = "Total Deductions"
    ws['J8'] = td
    ws['F9'] = "NET PAY"
    ws['G9'] = te - td
    ws['G8'] = te
    ws['F8'] = "Total Earnings"
    ws['G7'] = spal / 12
    ws['F7'] = "Special Allow"
    ws['G6'] = hra / 12
    ws['F6'] = "HRA"
    ws['G5'] = basic / 12
    ws['F5'] = "Basic"
    wb.save("Tax.xlsx")
